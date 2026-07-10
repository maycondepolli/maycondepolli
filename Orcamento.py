import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from PIL import Image, ImageTk
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

janela = tk.Tk()
janela.title("Orçamento Financeiro - Maycon Depolli dos Santos")

if sys.platform.startswith("win"):
    janela.state("zoomed")
else:
    janela.attributes("-zoomed", True)

# Nomes dos arquivos onde os dados serão salvos permanentemente
ARQUIVO_ENTRADAS = "entradas.txt"
ARQUIVO_SAIDAS = "saidas.txt"

# ==============================================================================
# --- VARIÁVEIS GLOBAIS PARA CONTROLE FINANCEIRO ---
# ==============================================================================
total_entradas_num = 0.0
total_saidas_num = 0.0
total_entradas_filtrado = 0.0
total_saidas_filtrado = 0.0
mes_export = "Todos"
ano_export = "Todos"

# ==============================================================================
# --- FUNÇÃO PARA EXTRAIR MÊS E ANO DA DATA ---
# ==============================================================================
def extrair_mes_ano(data_str):
    """Extrai o mês e ano de uma string de data no formato DD/MM/AAAA"""
    try:
        data_obj = datetime.strptime(data_str, "%d/%m/%Y")
        return data_obj.strftime("%m/%Y")  # Retorna MM/YYYY
    except:
        return None

def extrair_ano(data_str):
    """Extrai o ano de uma string de data no formato DD/MM/AAAA"""
    try:
        data_obj = datetime.strptime(data_str, "%d/%m/%Y")
        return data_obj.strftime("%Y")  # Retorna YYYY
    except:
        return None

# ==============================================================================
# --- FUNÇÃO PARA FILTRAR DADOS PARA EXPORTAÇÃO ---
# ==============================================================================
def filtrar_dados_para_export(mes_ref, ano_ref):
    """Filtra os dados de entrada e saída para exportação"""
    entradas_filtradas = []
    saidas_filtradas = []
    total_ent = 0.0
    total_sai = 0.0
    
    # Filtra entradas
    if os.path.exists(ARQUIVO_ENTRADAS):
        try:
            with open(ARQUIVO_ENTRADAS, "r", encoding="utf-8") as f:
                for linha in f:
                    partes = linha.strip().split(";")
                    if len(partes) == 5:
                        data, pago, categoria, valor_formatado, descricao = partes
                        mes_data = extrair_mes_ano(data)
                        ano_data = extrair_ano(data)
                        
                        # Verifica se deve incluir o registro
                        incluir = False
                        if mes_ref == "Todos" and ano_ref == "Todos":
                            incluir = True
                        elif mes_ref == "Todos" and ano_data == ano_ref:
                            incluir = True
                        elif mes_ref != "Todos" and ano_ref == "Todos" and mes_data and mes_data.startswith(mes_ref):
                            incluir = True
                        elif mes_ref != "Todos" and ano_ref != "Todos" and mes_data == f"{mes_ref}/{ano_ref}":
                            incluir = True
                        
                        if incluir:
                            entradas_filtradas.append((data, pago, categoria, valor_formatado, descricao))
                            valor_num = float(valor_formatado.replace(".", "").replace(",", "."))
                            total_ent += valor_num
        except Exception as e:
            print(f"Erro ao filtrar entradas: {e}")
    
    # Filtra saídas
    if os.path.exists(ARQUIVO_SAIDAS):
        try:
            with open(ARQUIVO_SAIDAS, "r", encoding="utf-8") as f:
                for linha in f:
                    partes = linha.strip().split(";")
                    if len(partes) == 5:
                        data, pago, categoria, valor_formatado, descricao = partes
                        mes_data = extrair_mes_ano(data)
                        ano_data = extrair_ano(data)
                        
                        # Verifica se deve incluir o registro
                        incluir = False
                        if mes_ref == "Todos" and ano_ref == "Todos":
                            incluir = True
                        elif mes_ref == "Todos" and ano_data == ano_ref:
                            incluir = True
                        elif mes_ref != "Todos" and ano_ref == "Todos" and mes_data and mes_data.startswith(mes_ref):
                            incluir = True
                        elif mes_ref != "Todos" and ano_ref != "Todos" and mes_data == f"{mes_ref}/{ano_ref}":
                            incluir = True
                        
                        if incluir:
                            saidas_filtradas.append((data, pago, categoria, valor_formatado, descricao))
                            valor_num = float(valor_formatado.replace(".", "").replace(",", "."))
                            total_sai += valor_num
        except Exception as e:
            print(f"Erro ao filtrar saídas: {e}")
    
    return entradas_filtradas, saidas_filtradas, total_ent, total_sai

# ==============================================================================
# --- FUNÇÃO PARA EXPORTAR DADOS PARA EXCEL ---
# ==============================================================================
def exportar_para_excel():
    """Exporta todos os dados filtrados para uma planilha Excel"""
    global mes_export, ano_export
    
    # Pega os valores dos comboboxes
    mes_export = combo_meses_export.get().split(" - ")[0] if combo_meses_export.get() != "Todos" else "Todos"
    ano_export = combo_anos_export.get()
    
    try:
        # Pergunta onde salvar o arquivo
        arquivo_salvar = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
            title="Salvar Planilha como"
        )
        
        if not arquivo_salvar:  # Usuário cancelou
            return
        
        # Filtra os dados
        entradas_filtradas, saidas_filtradas, total_ent, total_sai = filtrar_dados_para_export(mes_export, ano_export)
        
        # Verifica se há dados para exportar
        if not entradas_filtradas and not saidas_filtradas:
            messagebox.showwarning("Aviso", "Não há dados para o período selecionado!")
            return
        
        # Cria uma nova planilha
        wb = openpyxl.Workbook()
        
        # Remove a planilha padrão
        wb.remove(wb.active)
        
        # ========== PLANILHA DE ENTRADAS ==========
        ws_entradas = wb.create_sheet("Entradas")
        
        # Cabeçalhos
        cabecalhos_entradas = ["Data", "Pago", "Categoria", "Valor (R$)", "Descrição"]
        ws_entradas.append(cabecalhos_entradas)
        
        # Estiliza cabeçalhos
        for col in range(1, len(cabecalhos_entradas) + 1):
            celula = ws_entradas.cell(row=1, column=col)
            celula.font = Font(bold=True, color="FFFFFF")
            celula.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            celula.alignment = Alignment(horizontal="center", vertical="center")
        
        # Adiciona os dados de entrada filtrados
        for valores in entradas_filtradas:
            ws_entradas.append(valores)
        
        # Ajusta largura das colunas
        for col in range(1, len(cabecalhos_entradas) + 1):
            coluna_letra = get_column_letter(col)
            ws_entradas.column_dimensions[coluna_letra].width = 20
        
        # Adiciona total de entradas
        ultima_linha = ws_entradas.max_row + 1
        ws_entradas.cell(row=ultima_linha, column=4, value=f"Total: R$ {total_ent:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        ws_entradas.cell(row=ultima_linha, column=4).font = Font(bold=True)
        ws_entradas.cell(row=ultima_linha, column=4).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        
        # ========== PLANILHA DE SAÍDAS ==========
        ws_saidas = wb.create_sheet("Saídas")
        
        # Cabeçalhos
        cabecalhos_saidas = ["Data", "Pago", "Categoria", "Valor (R$)", "Descrição"]
        ws_saidas.append(cabecalhos_saidas)
        
        # Estiliza cabeçalhos
        for col in range(1, len(cabecalhos_saidas) + 1):
            celula = ws_saidas.cell(row=1, column=col)
            celula.font = Font(bold=True, color="FFFFFF")
            celula.fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
            celula.alignment = Alignment(horizontal="center", vertical="center")
        
        # Adiciona os dados de saída filtrados
        for valores in saidas_filtradas:
            ws_saidas.append(valores)
        
        # Ajusta largura das colunas
        for col in range(1, len(cabecalhos_saidas) + 1):
            coluna_letra = get_column_letter(col)
            ws_saidas.column_dimensions[coluna_letra].width = 20
        
        # Adiciona total de saídas
        ultima_linha = ws_saidas.max_row + 1
        ws_saidas.cell(row=ultima_linha, column=4, value=f"Total: R$ {total_sai:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        ws_saidas.cell(row=ultima_linha, column=4).font = Font(bold=True)
        ws_saidas.cell(row=ultima_linha, column=4).fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        
        # ========== PLANILHA DE RESUMO ==========
        ws_resumo = wb.create_sheet("Resumo")
        
        # Título
        ws_resumo.cell(row=1, column=1, value="RESUMO FINANCEIRO")
        ws_resumo.cell(row=1, column=1).font = Font(bold=True, size=16)
        ws_resumo.merge_cells('A1:E1')
        
        # Data do relatório
        ws_resumo.cell(row=2, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        ws_resumo.merge_cells('A2:E2')
        
        # Período filtrado
        if mes_export != "Todos" and ano_export != "Todos":
            nomes_meses = {
                '01': 'Janeiro', '02': 'Fevereiro', '03': 'Março', '04': 'Abril',
                '05': 'Maio', '06': 'Junho', '07': 'Julho', '08': 'Agosto',
                '09': 'Setembro', '10': 'Outubro', '11': 'Novembro', '12': 'Dezembro'
            }
            mes_nome = nomes_meses.get(mes_export, mes_export)
            ws_resumo.cell(row=3, column=1, value=f"Período: {mes_nome} de {ano_export}")
        else:
            ws_resumo.cell(row=3, column=1, value="Período: Todos os registros")
        ws_resumo.merge_cells('A3:E3')
        
        # Dados do resumo
        linha_atual = 5
        ws_resumo.cell(row=linha_atual, column=1, value="TOTAL DE ENTRADAS:")
        ws_resumo.cell(row=linha_atual, column=1).font = Font(bold=True)
        ws_resumo.cell(row=linha_atual, column=2, value=f"R$ {total_ent:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        ws_resumo.cell(row=linha_atual, column=2).font = Font(bold=True, color="4CAF50")
        
        linha_atual += 1
        ws_resumo.cell(row=linha_atual, column=1, value="TOTAL DE SAÍDAS:")
        ws_resumo.cell(row=linha_atual, column=1).font = Font(bold=True)
        ws_resumo.cell(row=linha_atual, column=2, value=f"R$ {total_sai:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        ws_resumo.cell(row=linha_atual, column=2).font = Font(bold=True, color="F44336")
        
        linha_atual += 1
        saldo = total_ent - total_sai
        ws_resumo.cell(row=linha_atual, column=1, value="SALDO FINAL:")
        ws_resumo.cell(row=linha_atual, column=1).font = Font(bold=True)
        ws_resumo.cell(row=linha_atual, column=2, value=f"R$ {saldo:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        ws_resumo.cell(row=linha_atual, column=2).font = Font(bold=True)
        if saldo >= 0:
            ws_resumo.cell(row=linha_atual, column=2).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        else:
            ws_resumo.cell(row=linha_atual, column=2).fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        
        # Quantidade de registros
        linha_atual += 2
        ws_resumo.cell(row=linha_atual, column=1, value="QUANTIDADE DE REGISTROS:")
        ws_resumo.cell(row=linha_atual, column=1).font = Font(bold=True)
        ws_resumo.cell(row=linha_atual, column=2, value=f"Entradas: {len(entradas_filtradas)} | Saídas: {len(saidas_filtradas)}")
        
        # Ajusta largura das colunas do resumo
        ws_resumo.column_dimensions['A'].width = 30
        ws_resumo.column_dimensions['B'].width = 25
        
        # ========== SALVA O ARQUIVO ==========
        wb.save(arquivo_salvar)
        messagebox.showinfo("Sucesso", f"Planilha exportada com sucesso!\n\nSalva em: {arquivo_salvar}")
        
    except ImportError:
        messagebox.showerror("Erro", "Biblioteca openpyxl não instalada!\n\nInstale com: pip install openpyxl")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao exportar planilha:\n\n{str(e)}")

# ==============================================================================
# --- FUNÇÃO PARA ATUALIZAR OS FILTROS DE EXPORTAÇÃO ---
# ==============================================================================
def atualizar_filtros_export():
    """Atualiza a lista de meses e anos disponíveis para exportação"""
    meses = set()
    anos = set()
    meses.add("Todos")
    anos.add("Todos")
    
    # Busca dados nas entradas
    if os.path.exists(ARQUIVO_ENTRADAS):
        try:
            with open(ARQUIVO_ENTRADAS, "r", encoding="utf-8") as f:
                for linha in f:
                    partes = linha.strip().split(";")
                    if len(partes) >= 1:
                        mes = extrair_mes_ano(partes[0])
                        ano = extrair_ano(partes[0])
                        if mes:
                            mes_num = mes.split('/')[0]
                            meses.add(mes_num)
                        if ano:
                            anos.add(ano)
        except:
            pass
    
    # Busca dados nas saídas
    if os.path.exists(ARQUIVO_SAIDAS):
        try:
            with open(ARQUIVO_SAIDAS, "r", encoding="utf-8") as f:
                for linha in f:
                    partes = linha.strip().split(";")
                    if len(partes) >= 1:
                        mes = extrair_mes_ano(partes[0])
                        ano = extrair_ano(partes[0])
                        if mes:
                            mes_num = mes.split('/')[0]
                            meses.add(mes_num)
                        if ano:
                            anos.add(ano)
        except:
            pass
    
    # Ordena os meses
    meses_validos = [m for m in meses if m != "Todos" and m.isdigit()]
    meses_ordenados = sorted(meses_validos, key=lambda x: int(x))
    
    nomes_meses = {
        '01': 'Janeiro', '02': 'Fevereiro', '03': 'Março', '04': 'Abril',
        '05': 'Maio', '06': 'Junho', '07': 'Julho', '08': 'Agosto',
        '09': 'Setembro', '10': 'Outubro', '11': 'Novembro', '12': 'Dezembro'
    }
    
    meses_com_nomes = []
    for m in meses_ordenados:
        meses_com_nomes.append(f"{m} - {nomes_meses.get(m, m)}")
    meses_com_nomes.insert(0, "Todos")
    
    anos_validos = [a for a in anos if a != "Todos" and a.isdigit()]
    anos_ordenados = sorted(anos_validos, key=lambda x: int(x))
    anos_ordenados.insert(0, "Todos")
    
    combo_meses_export['values'] = meses_com_nomes
    combo_anos_export['values'] = anos_ordenados
    
    if meses_com_nomes and len(meses_com_nomes) > 1:
        combo_meses_export.set("Todos")
    else:
        combo_meses_export.set("Todos")
    
    if anos_ordenados:
        combo_anos_export.set("Todos")

# ==============================================================================
# --- FUNÇÃO DE AUXÍLIO PARA ATUALIZAR O SALDO NA ABA 4 ---
# ==============================================================================
def atualizar_saldo_aba4():
    saldo_atual = total_entradas_num - total_saidas_num
    if saldo_atual >= 0:
        label_valor_saldo.config(fg="darkgreen")
    else:
        label_valor_saldo.config(fg="red")
    label_valor_saldo.config(text=f"R$ {saldo_atual:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

# ==============================================================================
# --- FUNÇÃO PARA CARREGAR DADOS SALVOS AO ABRIR O PROGRAMA ---
# ==============================================================================
def carregar_dados_salvos():
    global total_entradas_num, total_saidas_num
    
    # Zera os contadores antes de recontar os arquivos
    total_entradas_num = 0.0
    total_saidas_num = 0.0
    
    # Limpa as tabelas
    for item in tabela_entrada.get_children():
        tabela_entrada.delete(item)
    for item in tabela_saida.get_children():
        tabela_saida.delete(item)
    
    # Carrega as Entradas na tabela
    if os.path.exists(ARQUIVO_ENTRADAS):
        try:
            with open(ARQUIVO_ENTRADAS, "r", encoding="utf-8") as f:
                for linha in f:
                    partes = linha.strip().split(";")
                    if len(partes) == 5:
                        data, pago, categoria, valor_formatado, descricao = partes
                        tabela_entrada.insert("", "end", values=(data, pago, categoria, valor_formatado, descricao))
                        
                        # Converte a formatação monetária de volta para float puro do Python
                        valor_num = float(valor_formatado.replace(".", "").replace(",", "."))
                        total_entradas_num += valor_num
            label_soma_entradas.config(text=f"Total de Entradas: R$ {total_entradas_num:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        except Exception as e:
            print(f"Erro ao ler entradas salvas: {e}")

    # Carrega as Saídas na tabela
    if os.path.exists(ARQUIVO_SAIDAS):
        try:
            with open(ARQUIVO_SAIDAS, "r", encoding="utf-8") as f:
                for linha in f:
                    partes = linha.strip().split(";")
                    if len(partes) == 5:
                        data, pago, categoria, valor_formatado, descricao = partes
                        tabela_saida.insert("", "end", values=(data, pago, categoria, valor_formatado, descricao))
                        
                        # Converte a formatação monetária de volta para float puro do Python
                        valor_num = float(valor_formatado.replace(".", "").replace(",", "."))
                        total_saidas_num += valor_num
            label_soma_saidas.config(text=f"Total de Saídas: R$ {total_saidas_num:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        except Exception as e:
            print(f"Erro ao ler saídas salvas: {e}")

    # Atualiza o saldo geral na última aba
    atualizar_saldo_aba4()
    
    # Atualiza os filtros
    atualizar_filtros_entrada()
    atualizar_filtros_saida()
    atualizar_filtros_export()

# ==============================================================================
# --- FUNÇÕES PARA FILTRAR POR MÊS E ANO (INDEPENDENTES) ---
# ==============================================================================
def filtrar_entradas():
    """Filtra apenas a tabela de Entradas"""
    global total_entradas_num, total_entradas_filtrado
    
    mes_ref = combo_meses_entrada.get().split(" - ")[0] if combo_meses_entrada.get() != "Todos" else "Todos"
    ano_ref = combo_anos_entrada.get()
    
    # Limpa a tabela de entradas
    for item in tabela_entrada.get_children():
        tabela_entrada.delete(item)
    
    total_filtrado = 0.0
    
    # Lê o arquivo de entradas e filtra
    if os.path.exists(ARQUIVO_ENTRADAS):
        try:
            with open(ARQUIVO_ENTRADAS, "r", encoding="utf-8") as f:
                for linha in f:
                    partes = linha.strip().split(";")
                    if len(partes) == 5:
                        data, pago, categoria, valor_formatado, descricao = partes
                        mes_data = extrair_mes_ano(data)
                        ano_data = extrair_ano(data)
                        
                        # Verifica se deve incluir o registro
                        incluir = False
                        if mes_ref == "Todos" and ano_ref == "Todos":
                            incluir = True
                        elif mes_ref == "Todos" and ano_data == ano_ref:
                            incluir = True
                        elif mes_ref != "Todos" and ano_ref == "Todos" and mes_data and mes_data.startswith(mes_ref):
                            incluir = True
                        elif mes_ref != "Todos" and ano_ref != "Todos" and mes_data == f"{mes_ref}/{ano_ref}":
                            incluir = True
                        
                        if incluir:
                            tabela_entrada.insert("", "end", values=(data, pago, categoria, valor_formatado, descricao))
                            valor_num = float(valor_formatado.replace(".", "").replace(",", "."))
                            total_filtrado += valor_num
        except Exception as e:
            print(f"Erro ao filtrar entradas: {e}")
    
    total_entradas_num = total_filtrado
    label_soma_entradas.config(text=f"Total de Entradas: R$ {total_entradas_num:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    
    # Atualiza o saldo
    atualizar_saldo_aba4()

def filtrar_saidas():
    """Filtra apenas a tabela de Saídas"""
    global total_saidas_num, total_saidas_filtrado
    
    mes_ref = combo_meses_saida.get().split(" - ")[0] if combo_meses_saida.get() != "Todos" else "Todos"
    ano_ref = combo_anos_saida.get()
    
    # Limpa a tabela de saídas
    for item in tabela_saida.get_children():
        tabela_saida.delete(item)
    
    total_filtrado = 0.0
    
    # Lê o arquivo de saídas e filtra
    if os.path.exists(ARQUIVO_SAIDAS):
        try:
            with open(ARQUIVO_SAIDAS, "r", encoding="utf-8") as f:
                for linha in f:
                    partes = linha.strip().split(";")
                    if len(partes) == 5:
                        data, pago, categoria, valor_formatado, descricao = partes
                        mes_data = extrair_mes_ano(data)
                        ano_data = extrair_ano(data)
                        
                        # Verifica se deve incluir o registro
                        incluir = False
                        if mes_ref == "Todos" and ano_ref == "Todos":
                            incluir = True
                        elif mes_ref == "Todos" and ano_data == ano_ref:
                            incluir = True
                        elif mes_ref != "Todos" and ano_ref == "Todos" and mes_data and mes_data.startswith(mes_ref):
                            incluir = True
                        elif mes_ref != "Todos" and ano_ref != "Todos" and mes_data == f"{mes_ref}/{ano_ref}":
                            incluir = True
                        
                        if incluir:
                            tabela_saida.insert("", "end", values=(data, pago, categoria, valor_formatado, descricao))
                            valor_num = float(valor_formatado.replace(".", "").replace(",", "."))
                            total_filtrado += valor_num
        except Exception as e:
            print(f"Erro ao filtrar saídas: {e}")
    
    total_saidas_num = total_filtrado
    label_soma_saidas.config(text=f"Total de Saídas: R$ {total_saidas_num:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    
    # Atualiza o saldo
    atualizar_saldo_aba4()

def mostrar_todas_entradas():
    """Mostra todas as entradas sem filtro"""
    global total_entradas_num
    
    # Limpa a tabela
    for item in tabela_entrada.get_children():
        tabela_entrada.delete(item)
    
    total_entradas_num = 0.0
    
    if os.path.exists(ARQUIVO_ENTRADAS):
        try:
            with open(ARQUIVO_ENTRADAS, "r", encoding="utf-8") as f:
                for linha in f:
                    partes = linha.strip().split(";")
                    if len(partes) == 5:
                        data, pago, categoria, valor_formatado, descricao = partes
                        tabela_entrada.insert("", "end", values=(data, pago, categoria, valor_formatado, descricao))
                        valor_num = float(valor_formatado.replace(".", "").replace(",", "."))
                        total_entradas_num += valor_num
        except Exception as e:
            print(f"Erro ao ler entradas: {e}")
    
    label_soma_entradas.config(text=f"Total de Entradas: R$ {total_entradas_num:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    
    # Atualiza os filtros
    atualizar_filtros_entrada()
    atualizar_saldo_aba4()

def mostrar_todas_saidas():
    """Mostra todas as saídas sem filtro"""
    global total_saidas_num
    
    # Limpa a tabela
    for item in tabela_saida.get_children():
        tabela_saida.delete(item)
    
    total_saidas_num = 0.0
    
    if os.path.exists(ARQUIVO_SAIDAS):
        try:
            with open(ARQUIVO_SAIDAS, "r", encoding="utf-8") as f:
                for linha in f:
                    partes = linha.strip().split(";")
                    if len(partes) == 5:
                        data, pago, categoria, valor_formatado, descricao = partes
                        tabela_saida.insert("", "end", values=(data, pago, categoria, valor_formatado, descricao))
                        valor_num = float(valor_formatado.replace(".", "").replace(",", "."))
                        total_saidas_num += valor_num
        except Exception as e:
            print(f"Erro ao ler saídas: {e}")
    
    label_soma_saidas.config(text=f"Total de Saídas: R$ {total_saidas_num:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    
    # Atualiza os filtros
    atualizar_filtros_saida()
    atualizar_saldo_aba4()

# ==============================================================================
# --- FUNÇÕES PARA ATUALIZAR OS FILTROS ---
# ==============================================================================
def atualizar_filtros_entrada():
    """Atualiza a lista de meses e anos disponíveis para entradas"""
    meses = set()
    anos = set()
    meses.add("Todos")
    anos.add("Todos")
    
    if os.path.exists(ARQUIVO_ENTRADAS):
        try:
            with open(ARQUIVO_ENTRADAS, "r", encoding="utf-8") as f:
                for linha in f:
                    partes = linha.strip().split(";")
                    if len(partes) >= 1:
                        mes = extrair_mes_ano(partes[0])
                        ano = extrair_ano(partes[0])
                        if mes:
                            mes_num = mes.split('/')[0]
                            meses.add(mes_num)
                        if ano:
                            anos.add(ano)
        except:
            pass
    
    # Ordena os meses
    meses_validos = [m for m in meses if m != "Todos" and m.isdigit()]
    meses_ordenados = sorted(meses_validos, key=lambda x: int(x))
    
    nomes_meses = {
        '01': 'Janeiro', '02': 'Fevereiro', '03': 'Março', '04': 'Abril',
        '05': 'Maio', '06': 'Junho', '07': 'Julho', '08': 'Agosto',
        '09': 'Setembro', '10': 'Outubro', '11': 'Novembro', '12': 'Dezembro'
    }
    
    meses_com_nomes = []
    for m in meses_ordenados:
        meses_com_nomes.append(f"{m} - {nomes_meses.get(m, m)}")
    meses_com_nomes.insert(0, "Todos")
    
    anos_validos = [a for a in anos if a != "Todos" and a.isdigit()]
    anos_ordenados = sorted(anos_validos, key=lambda x: int(x))
    anos_ordenados.insert(0, "Todos")
    
    combo_meses_entrada['values'] = meses_com_nomes
    combo_anos_entrada['values'] = anos_ordenados
    
    if meses_com_nomes and len(meses_com_nomes) > 1:
        combo_meses_entrada.set("Todos")
    else:
        combo_meses_entrada.set("Todos")
    
    if anos_ordenados:
        combo_anos_entrada.set("Todos")

def atualizar_filtros_saida():
    """Atualiza a lista de meses e anos disponíveis para saídas"""
    meses = set()
    anos = set()
    meses.add("Todos")
    anos.add("Todos")
    
    if os.path.exists(ARQUIVO_SAIDAS):
        try:
            with open(ARQUIVO_SAIDAS, "r", encoding="utf-8") as f:
                for linha in f:
                    partes = linha.strip().split(";")
                    if len(partes) >= 1:
                        mes = extrair_mes_ano(partes[0])
                        ano = extrair_ano(partes[0])
                        if mes:
                            mes_num = mes.split('/')[0]
                            meses.add(mes_num)
                        if ano:
                            anos.add(ano)
        except:
            pass
    
    # Ordena os meses
    meses_validos = [m for m in meses if m != "Todos" and m.isdigit()]
    meses_ordenados = sorted(meses_validos, key=lambda x: int(x))
    
    nomes_meses = {
        '01': 'Janeiro', '02': 'Fevereiro', '03': 'Março', '04': 'Abril',
        '05': 'Maio', '06': 'Junho', '07': 'Julho', '08': 'Agosto',
        '09': 'Setembro', '10': 'Outubro', '11': 'Novembro', '12': 'Dezembro'
    }
    
    meses_com_nomes = []
    for m in meses_ordenados:
        meses_com_nomes.append(f"{m} - {nomes_meses.get(m, m)}")
    meses_com_nomes.insert(0, "Todos")
    
    anos_validos = [a for a in anos if a != "Todos" and a.isdigit()]
    anos_ordenados = sorted(anos_validos, key=lambda x: int(x))
    anos_ordenados.insert(0, "Todos")
    
    combo_meses_saida['values'] = meses_com_nomes
    combo_anos_saida['values'] = anos_ordenados
    
    if meses_com_nomes and len(meses_com_nomes) > 1:
        combo_meses_saida.set("Todos")
    else:
        combo_meses_saida.set("Todos")
    
    if anos_ordenados:
        combo_anos_saida.set("Todos")

# ==============================================================================
# --- FUNÇÃO PARA EDITAR REGISTRO ---
# ==============================================================================
def editar_registro(tabela, arquivo, tipo):
    selecionado = tabela.selection()
    if not selecionado:
        messagebox.showwarning("Aviso", "Selecione um registro para editar.")
        return
    
    # Pega os valores atuais
    valores = tabela.item(selecionado[0])['values']
    
    # Cria uma janela de edição
    janela_edicao = tk.Toplevel(janela)
    janela_edicao.title("Editar Registro")
    janela_edicao.geometry("500x450")
    janela_edicao.transient(janela)
    janela_edicao.grab_set()
    
    # Campos de edição
    tk.Label(janela_edicao, text="Data (DD/MM/AAAA):").grid(row=0, column=0, padx=10, pady=5, sticky="w")
    entry_data = ttk.Entry(janela_edicao, width=30)
    entry_data.grid(row=0, column=1, padx=10, pady=5)
    entry_data.insert(0, valores[0])
    
    tk.Label(janela_edicao, text="Pago (em branco = Entrada):").grid(row=1, column=0, padx=10, pady=5, sticky="w")
    entry_pago = ttk.Entry(janela_edicao, width=30)
    entry_pago.grid(row=1, column=1, padx=10, pady=5)
    entry_pago.insert(0, valores[1])
    
    tk.Label(janela_edicao, text="Categoria:").grid(row=2, column=0, padx=10, pady=5, sticky="w")
    entry_categoria = ttk.Entry(janela_edicao, width=30)
    entry_categoria.grid(row=2, column=1, padx=10, pady=5)
    entry_categoria.insert(0, valores[2])
    
    tk.Label(janela_edicao, text="Valor:").grid(row=3, column=0, padx=10, pady=5, sticky="w")
    entry_valor = ttk.Entry(janela_edicao, width=30)
    entry_valor.grid(row=3, column=1, padx=10, pady=5)
    entry_valor.insert(0, valores[3])
    
    tk.Label(janela_edicao, text="Descrição:").grid(row=4, column=0, padx=10, pady=5, sticky="nw")
    entry_descricao = tk.Text(janela_edicao, width=30, height=8)
    entry_descricao.grid(row=4, column=1, padx=10, pady=5)
    entry_descricao.insert("1.0", valores[4])
    
    def salvar_edicao():
        novo_data = entry_data.get()
        novo_pago = entry_pago.get()
        novo_categoria = entry_categoria.get()
        novo_valor = entry_valor.get()
        novo_descricao = entry_descricao.get("1.0", "end-1c")
        
        # Atualiza na tabela
        tabela.item(selecionado[0], values=(novo_data, novo_pago, novo_categoria, novo_valor, novo_descricao))
        
        # Atualiza no arquivo
        atualizar_arquivo(tabela, arquivo)
        
        # Recarrega os dados
        carregar_dados_salvos()
        
        janela_edicao.destroy()
        messagebox.showinfo("Sucesso", "Registro editado com sucesso!")
    
    btn_salvar = tk.Button(janela_edicao, text="Salvar Alterações", command=salvar_edicao, bg="green", fg="white", font=("Arial", 10, "bold"))
    btn_salvar.grid(row=5, column=0, columnspan=2, pady=20)

# ==============================================================================
# --- FUNÇÃO PARA EXCLUIR REGISTRO ---
# ==============================================================================
def excluir_registro(tabela, arquivo):
    selecionado = tabela.selection()
    if not selecionado:
        messagebox.showwarning("Aviso", "Selecione um registro para excluir.")
        return
    
    if messagebox.askyesno("Confirmar Exclusão", "Tem certeza que deseja excluir este registro?"):
        tabela.delete(selecionado[0])
        atualizar_arquivo(tabela, arquivo)
        carregar_dados_salvos()
        messagebox.showinfo("Sucesso", "Registro excluído com sucesso!")

# ==============================================================================
# --- FUNÇÃO PARA ATUALIZAR ARQUIVO ---
# ==============================================================================
def atualizar_arquivo(tabela, arquivo):
    try:
        with open(arquivo, "w", encoding="utf-8") as f:
            for item in tabela.get_children():
                valores = tabela.item(item)['values']
                linha = ";".join(str(v) for v in valores)
                f.write(linha + "\n")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao atualizar arquivo: {e}")

# ==============================================================================
# --- FUNÇÃO DE AÇÃO DO BOTÃO SALVAR (MODIFICADA) ---
# ==============================================================================
def salvar_dados():
    global total_entradas_num, total_saidas_num
    
    # 1. Coleta as informações digitadas nos campos da Aba 1
    data = entrada_data.get()
    pago = entrada_pago.get().strip()  
    categoria = entrada_categoria.get()
    valor_texto = entrada_valor.get()
    descricao = entrada_descricao_longa.get("1.0", "end-1c")
    
    # Validação básica
    if not data or not categoria or not valor_texto:
        messagebox.showwarning("Aviso", "Preencha todos os campos obrigatórios (Data, Categoria e Valor)!")
        return
    
    # Valida formato da data
    try:
        datetime.strptime(data, "%d/%m/%Y")
    except:
        messagebox.showerror("Erro", "Data inválida! Use o formato DD/MM/AAAA")
        return
    
    descricao_limpa = descricao.replace("\n", " ").replace(";", " ")
    
    # 2. Tratamento e conversão segura do valor para número decimal
    try:
        valor_limpo = valor_texto.replace("R$", "").replace(".", "").replace(",", ".").strip()
        valor_num = float(valor_limpo)
    except ValueError:
        messagebox.showerror("Erro", "Valor inválido! Use formato como 1500,00 ou 1.500,00")
        return
        
    valor_formatado = f"{valor_num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    # 3. NOVA LÓGICA: Se campo Pago estiver vazio -> ENTRADA, se preenchido -> SAÍDA
    if pago == "":  # Campo vazio - vai para ENTRADA
        # Adiciona à tabela de entradas
        tabela_entrada.insert("", "end", values=(data, pago, categoria, valor_formatado, descricao_limpa))
        
        with open(ARQUIVO_ENTRADAS, "a", encoding="utf-8") as f:
            f.write(f"{data};{pago};{categoria};{valor_formatado};{descricao_limpa}\n")
        
        total_entradas_num += valor_num
        label_soma_entradas.config(text=f"Total de Entradas: R$ {total_entradas_num:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    else:  # Campo preenchido (qualquer texto) - vai para SAÍDA
        # Adiciona à tabela de saídas
        tabela_saida.insert("", "end", values=(data, pago, categoria, valor_formatado, descricao_limpa))
        
        with open(ARQUIVO_SAIDAS, "a", encoding="utf-8") as f:
            f.write(f"{data};{pago};{categoria};{valor_formatado};{descricao_limpa}\n")
        
        total_saidas_num += valor_num
        label_soma_saidas.config(text=f"Total de Saídas: R$ {total_saidas_num:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

    # 4. Atualiza o saldo geral na última aba
    atualizar_saldo_aba4()

    # 5. Limpa os campos da Aba 1 após salvar para o próximo preenchimento
    entrada_data.delete(0, tk.END)
    entrada_pago.delete(0, tk.END)
    entrada_categoria.delete(0, tk.END)
    entrada_valor.delete(0, tk.END)
    entrada_descricao_longa.delete("1.0", tk.END)
    
    # 6. Atualiza os filtros
    atualizar_filtros_entrada()
    atualizar_filtros_saida()
    atualizar_filtros_export()
    
    messagebox.showinfo("Sucesso", "Registro salvo com sucesso!")

# ==============================================================================

# Criação do Notebook (Gerenciador de Abas)
notebook = ttk.Notebook(janela)

# --- Aba 1 ---
aba1 = ttk.Frame(notebook)
notebook.add(aba1, text="Valores a serem Cadastrados")

# --- Aba 2 ---
aba2 = ttk.Frame(notebook)
notebook.add(aba2, text="Entrada ")

# --- Aba 3 ---
aba3 = ttk.Frame(notebook)
notebook.add(aba3, text="Saida")

aba4 = ttk.Frame(notebook)
notebook.add(aba4, text="Saldo Atual")

# Empacota o notebook para que ele apareça na janela
notebook.pack(expand=True, fill="both")

# --- CRIANDO UM RODAPÉ NA JANELA PRINCIPAL ---
rodape = tk.Frame(janela)
rodape.pack(side="bottom", fill="x", padx=20, pady=10)

# Carregamento da imagem com tamanho maior
bg_image = None
try:
    diretorio_atual = os.path.dirname(os.path.abspath(__file__))
    caminho_imagem = os.path.join(diretorio_atual, "Imagens", "Maycon.png")
    if os.path.exists(caminho_imagem):
        imagem_pil = Image.open(caminho_imagem)
        imagem_pil = imagem_pil.resize((150, 150), Image.Resampling.LANCZOS)
        bg_image = ImageTk.PhotoImage(imagem_pil)
        print("Imagem carregada com sucesso!")
    else:
        print(f"Arquivo de imagem não encontrado em: {caminho_imagem}")
except Exception as e:
    print(f"Erro ao carregar a imagem: {e}")

# ==============================================================================
# --- CENTRALIZAÇÃO DO RODAPÉ COM PACK E LOGO MAIOR ---
# ==============================================================================
if bg_image:
    container_rodape = tk.Frame(rodape)
    container_rodape.pack(anchor="center", pady=5)
    
    label_logo = tk.Label(container_rodape, image=bg_image)
    label_logo.pack(pady=5)
    label_logo.image = bg_image
    
    label_creditos = tk.Label(
        container_rodape,
        text="Desenvolvido por: Maycon Depolli dos Santos",
        font=("Arial", 14, "bold"),
        fg="blue",
    )
    label_creditos.pack(pady=5)
else:
    label_Quem_fez = tk.Label(
        rodape,
        text="Desenvolvido por: Maycon Depolli dos Santos",
        font=("Arial", 14, "bold"),
        fg="blue",
    )
    label_Quem_fez.pack(anchor="center", pady=5)

# ==============================================================================
# --- ADICIONANDO CAMPOS EM CADA ABA ---
# ==============================================================================

# --- Campos da Aba 1 (Valores a serem Cadastrados) ---
# Linha 1: Data
label_data = tk.Label(aba1, text="Data (DD/MM/AAAA):", font=("Arial", 11))
label_data.place(relx=0.36, y=30, anchor="w") 

entrada_data = ttk.Entry(aba1, width=30)
entrada_data.place(relx=0.50, y=30, anchor="w")

# Linha 2: Pago (modificado)
label_pago = tk.Label(aba1, text="Pago (em branco = Entrada):", font=("Arial", 11))
label_pago.place(relx=0.36, y=80, anchor="w")

entrada_pago = ttk.Entry(aba1, width=30)
entrada_pago.place(relx=0.50, y=80, anchor="w")

# Linha 3: Categoria
label_categoria = tk.Label(aba1, text="Categoria:", font=("Arial", 11))
label_categoria.place(relx=0.36, y=130, anchor="w")

entrada_categoria = ttk.Entry(aba1, width=30)
entrada_categoria.place(relx=0.50, y=130, anchor="w")

# Linha 4: Valor
label_valor = tk.Label(aba1, text="Valor:", font=("Arial", 11))
label_valor.place(relx=0.36, y=180, anchor="w")

entrada_valor = ttk.Entry(aba1, width=30)
entrada_valor.place(relx=0.50, y=180, anchor="w")

# Linha 5: Descrição do Item
label_descricao_longa = tk.Label(aba1, text="Descrição do Item:", font=("Arial", 11))
label_descricao_longa.place(relx=0.36, y=230, anchor="nw")

entrada_descricao_longa = tk.Text(aba1, width=26, height=10, font=("Arial", 10), bd=1, relief="solid")
entrada_descricao_longa.place(relx=0.50, y=230, anchor="nw")

botao_salvar = tk.Button(aba1,text="Salvar",command=salvar_dados,width=15,height=2,font=("Arial", 10, "bold"), bg="#2196F3", fg="white")
botao_salvar.place(relx=0.50, y=420, anchor="n")

# --- Aba 2: Entrada com Filtros Independentes ---
# Frame para filtros
frame_filtros_entrada = tk.Frame(aba2, bg="#f0f0f0", relief="sunken", bd=1)
frame_filtros_entrada.place(x=0, y=0, relwidth=1.0, height=40)

tk.Label(frame_filtros_entrada, text="Filtrar por Mês:", font=("Arial", 10), bg="#f0f0f0").place(x=10, y=10)
combo_meses_entrada = ttk.Combobox(frame_filtros_entrada, width=20, state="readonly")
combo_meses_entrada.place(x=110, y=8)

tk.Label(frame_filtros_entrada, text="Ano:", font=("Arial", 10), bg="#f0f0f0").place(x=280, y=10)
combo_anos_entrada = ttk.Combobox(frame_filtros_entrada, width=10, state="readonly")
combo_anos_entrada.place(x=320, y=8)

btn_aplicar_filtro_entrada = tk.Button(
    frame_filtros_entrada,
    text="🔍 Aplicar Filtro",
    command=filtrar_entradas,
    bg="#4CAF50",
    fg="white",
    font=("Arial", 9, "bold"),
    padx=10
)
btn_aplicar_filtro_entrada.place(x=430, y=5)

btn_mostrar_todas_entradas = tk.Button(
    frame_filtros_entrada,
    text="📋 Mostrar Todas",
    command=mostrar_todas_entradas,
    bg="#FF9800",
    fg="white",
    font=("Arial", 9, "bold"),
    padx=10
)
btn_mostrar_todas_entradas.place(x=550, y=5)

# Tabela de Entradas
frame_tabela_entrada = tk.Frame(aba2)
frame_tabela_entrada.place(x=0, y=45, relwidth=1.0, relheight=0.65)

colunas_entrada = ("data", "pago", "categoria", "valor", "descricao")
tabela_entrada = ttk.Treeview(frame_tabela_entrada, columns=colunas_entrada, show="headings")
tabela_entrada.heading("data", text="Data")
tabela_entrada.heading("pago", text="Pago")
tabela_entrada.heading("categoria", text="Categoria")
tabela_entrada.heading("valor", text="Valor (R$)")
tabela_entrada.heading("descricao", text="Descrição do Item")
tabela_entrada.column("data", minwidth=100, stretch=True, anchor="center")
tabela_entrada.column("pago", minwidth=80, stretch=True, anchor="center")
tabela_entrada.column("categoria", minwidth=140, stretch=True, anchor="w")
tabela_entrada.column("valor", minwidth=120, stretch=True, anchor="e")
tabela_entrada.column("descricao", minwidth=250, stretch=True, anchor="w")

scroll_entrada = ttk.Scrollbar(frame_tabela_entrada, orient="vertical", command=tabela_entrada.yview)
tabela_entrada.configure(yscrollcommand=scroll_entrada.set)
tabela_entrada.pack(side="left", fill="both", expand=True)
scroll_entrada.pack(side="right", fill="y")

# Frame para botões da aba Entrada
frame_botoes_entrada = tk.Frame(aba2)
frame_botoes_entrada.place(relx=0.5, rely=0.85, anchor="center")

btn_editar_entrada = tk.Button(
    frame_botoes_entrada, 
    text="✏️ Editar", 
    command=lambda: editar_registro(tabela_entrada, ARQUIVO_ENTRADAS, "entrada"),
    bg="#4CAF50",
    fg="white",
    font=("Arial", 10, "bold"),
    padx=25,
    pady=8,
    width=10
)
btn_editar_entrada.pack(side="left", padx=15)

btn_excluir_entrada = tk.Button(
    frame_botoes_entrada, 
    text="🗑️ Excluir", 
    command=lambda: excluir_registro(tabela_entrada, ARQUIVO_ENTRADAS),
    bg="#f44336",
    fg="white",
    font=("Arial", 10, "bold"),
    padx=25,
    pady=8,
    width=10
)
btn_excluir_entrada.pack(side="left", padx=15)

label_soma_entradas = tk.Label(aba2,text="Total de Entradas: R$ 0,00",font=("Arial", 14, "bold"),fg="green")
label_soma_entradas.place(relx=0.89, rely=0.85, anchor="e")

# --- Aba 3: Saída com Filtros Independentes ---
# Frame para filtros
frame_filtros_saida = tk.Frame(aba3, bg="#f0f0f0", relief="sunken", bd=1)
frame_filtros_saida.place(x=0, y=0, relwidth=1.0, height=40)

tk.Label(frame_filtros_saida, text="Filtrar por Mês:", font=("Arial", 10), bg="#f0f0f0").place(x=10, y=10)
combo_meses_saida = ttk.Combobox(frame_filtros_saida, width=20, state="readonly")
combo_meses_saida.place(x=110, y=8)

tk.Label(frame_filtros_saida, text="Ano:", font=("Arial", 10), bg="#f0f0f0").place(x=280, y=10)
combo_anos_saida = ttk.Combobox(frame_filtros_saida, width=10, state="readonly")
combo_anos_saida.place(x=320, y=8)

btn_aplicar_filtro_saida = tk.Button(
    frame_filtros_saida,
    text="🔍 Aplicar Filtro",
    command=filtrar_saidas,
    bg="#4CAF50",
    fg="white",
    font=("Arial", 9, "bold"),
    padx=10
)
btn_aplicar_filtro_saida.place(x=430, y=5)

btn_mostrar_todas_saidas = tk.Button(
    frame_filtros_saida,
    text="📋 Mostrar Todas",
    command=mostrar_todas_saidas,
    bg="#FF9800",
    fg="white",
    font=("Arial", 9, "bold"),
    padx=10
)
btn_mostrar_todas_saidas.place(x=550, y=5)

# Tabela de Saídas
frame_tabela_saida = tk.Frame(aba3)
frame_tabela_saida.place(x=0, y=45, relwidth=1.0, relheight=0.65)

colunas_saida = ("data", "pago", "categoria", "valor", "descricao")
tabela_saida = ttk.Treeview(frame_tabela_saida, columns=colunas_saida, show="headings")
tabela_saida.heading("data", text="Data")
tabela_saida.heading("pago", text="Pago")
tabela_saida.heading("categoria", text="Categoria")
tabela_saida.heading("valor", text="Valor (R$)")
tabela_saida.heading("descricao", text="Descrição do Item")
tabela_saida.column("data", minwidth=100, stretch=True, anchor="center")
tabela_saida.column("pago", minwidth=80, stretch=True, anchor="center")
tabela_saida.column("categoria", minwidth=140, stretch=True, anchor="w")
tabela_saida.column("valor", minwidth=120, stretch=True, anchor="e")
tabela_saida.column("descricao", minwidth=250, stretch=True, anchor="w")

scroll_saida = ttk.Scrollbar(frame_tabela_saida, orient="vertical", command=tabela_saida.yview)
tabela_saida.configure(yscrollcommand=scroll_saida.set)
tabela_saida.pack(side="left", fill="both", expand=True)
scroll_saida.pack(side="right", fill="y")

# Frame para botões da aba Saída
frame_botoes_saida = tk.Frame(aba3)
frame_botoes_saida.place(relx=0.5, rely=0.85, anchor="center")

btn_editar_saida = tk.Button(
    frame_botoes_saida, 
    text="✏️ Editar", 
    command=lambda: editar_registro(tabela_saida, ARQUIVO_SAIDAS, "saida"),
    bg="#4CAF50",
    fg="white",
    font=("Arial", 10, "bold"),
    padx=25,
    pady=8,
    width=10
)
btn_editar_saida.pack(side="left", padx=15)

btn_excluir_saida = tk.Button(
    frame_botoes_saida, 
    text="🗑️ Excluir", 
    command=lambda: excluir_registro(tabela_saida, ARQUIVO_SAIDAS),
    bg="#f44336",
    fg="white",
    font=("Arial", 10, "bold"),
    padx=25,
    pady=8,
    width=10
)
btn_excluir_saida.pack(side="left", padx=15)

label_soma_saidas = tk.Label(aba3,text="Total de Saídas: R$ 0,00",font=("Arial", 14, "bold"),fg="red")
label_soma_saidas.place(relx=0.89, rely=0.85, anchor="e")

# --- Aba 4: Saldo com Filtros para Exportação ---
# Título
label_campo_aba4 = tk.Label(aba4, text="Seu Saldo Disponível:", font=("Arial", 14, "bold"))
label_campo_aba4.place(relx=0.5, y=30, anchor="center")

label_valor_saldo = tk.Label(aba4,text="R$ 0,00",font=("Arial", 24, "bold"),fg="darkgreen")
label_valor_saldo.place(relx=0.5, y=70, anchor="center")

# Frame para filtros de exportação
frame_filtros_export = tk.Frame(aba4, bg="#f0f0f0", relief="sunken", bd=1)
frame_filtros_export.place(relx=0.5, y=130, anchor="center", width=500, height=50)

tk.Label(frame_filtros_export, text="Selecionar período para exportar:", font=("Arial", 10, "bold"), bg="#f0f0f0").place(x=10, y=5)

tk.Label(frame_filtros_export, text="Mês:", font=("Arial", 9), bg="#f0f0f0").place(x=10, y=28)
combo_meses_export = ttk.Combobox(frame_filtros_export, width=15, state="readonly")
combo_meses_export.place(x=50, y=25)

tk.Label(frame_filtros_export, text="Ano:", font=("Arial", 9), bg="#f0f0f0").place(x=200, y=28)
combo_anos_export = ttk.Combobox(frame_filtros_export, width=10, state="readonly")
combo_anos_export.place(x=240, y=25)

# Botão para exportar para Excel
btn_exportar_excel = tk.Button(
    aba4,
    text="📊 Exportar para Excel",
    command=exportar_para_excel,
    bg="#4CAF50",
    fg="white",
    font=("Arial", 12, "bold"),
    padx=30,
    pady=10
)
btn_exportar_excel.place(relx=0.5, y=210, anchor="center")

# Label informativo
label_info_export = tk.Label(
    aba4,
    text="Selecione o período desejado e clique em Exportar para Excel",
    font=("Arial", 9),
    fg="gray"
)
label_info_export.place(relx=0.5, y=260, anchor="center")

# Carrega os dados salvos ao iniciar
carregar_dados_salvos()

janela.mainloop()
